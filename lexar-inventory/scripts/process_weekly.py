#!/usr/bin/env python3
"""
Lexar 客户进销存周报数据处理脚本
每周一自动运行：读取飞书数据源 → 生成库存和销售上传文件
"""
import subprocess, json, csv, io, sys, os
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font

FEISHU_URL = "https://xxx0an5xw1s.feishu.cn/sheets/IjPGsnJ1ohjMzWtSaITcf3ssnQf"
SALE_SHEET_ID = "ULyVmO"
OUTPUT_DIR = "/Users/Zhuanz/Documents/lexar进销存/weekly_upload"

WAREHOUSE_RULES = {
    'online': lambda w: 'AMAZON' in w.upper(),
    'NL': lambda w: w == '7117.荷兰EI完税仓',
    'HK': lambda w: w in ('7436.中山保税HKYS成品仓', '7437.中山保税GPSR成品仓'),
}


def run_lark_cli(args):
    cmd = ["lark-cli", "sheets"] + args
    result = subprocess.run(cmd, capture_output=True, text=True)
    output = result.stdout
    idx = output.find('{')
    if idx < 0:
        print(f"ERROR: lark-cli returned no JSON: {result.stderr}", file=sys.stderr)
        sys.exit(1)
    return json.loads(output[idx:])


def parse_csv_rows(annotated_csv):
    lines = annotated_csv.strip().split('\n')
    header_text = lines[0].split('] ', 1)[1]
    reader = csv.reader(io.StringIO(header_text))
    headers = next(reader)
    rows = []
    for line in lines[1:]:
        row_text = line.split('] ', 1)[1]
        reader = csv.reader(io.StringIO(row_text))
        row = next(reader)
        rows.append(dict(zip(headers, row)))
    return headers, rows


def get_workbook_info():
    data = run_lark_cli(["+workbook-info", "--url", FEISHU_URL])
    return data['data']['sheets']


def find_stock_sheet(sheets):
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    target_suffix = monday.strftime("%m%d")

    for s in sheets:
        name = s['sheet_name']
        if name.startswith('STOCK-') and name.endswith(target_suffix):
            return s
    for s in sheets:
        name = s['sheet_name']
        if name.startswith('STOCK-'):
            return s
    return None


def process_inventory(stock_sheet):
    sheet_id = stock_sheet['sheet_id']
    row_count = stock_sheet['row_count']
    col_count = stock_sheet['column_count']

    max_col = chr(ord('A') + min(col_count - 1, 25))
    data = run_lark_cli([
        "+csv-get", "--url", FEISHU_URL,
        "--sheet-id", sheet_id,
        "--range", f"A1:{max_col}{row_count}"
    ])

    headers, rows = parse_csv_rows(data['data']['annotated_csv'])

    inventory = defaultdict(lambda: {'HK': 0, 'NL': 0, 'online': 0})
    for row in rows:
        warehouse = row.get('仓库', '').strip()
        pn = row.get('产品P/N', '').strip()
        qty_str = row.get('现有量', '').strip()
        if not pn:
            continue
        try:
            qty = int(float(qty_str)) if qty_str else 0
        except ValueError:
            qty = 0

        for key, rule in WAREHOUSE_RULES.items():
            if rule(warehouse):
                inventory[pn][key] += qty
                break

    wb = Workbook()
    ws = wb.active
    ws.title = '期末库存上传模版'
    ws.append(['COMPANY', 'PN', 'lexar HK库存', 'lexar NL库存', 'lexar online库存', 'total库存'])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    row_idx = 2
    for pn in sorted(inventory.keys()):
        d = inventory[pn]
        hk = d['HK'] if d['HK'] else None
        nl = d['NL'] if d['NL'] else None
        online = d['online'] if d['online'] else None
        ws.append(['LEXAR', pn, hk, nl, online])
        ws.cell(row=row_idx, column=6).value = f'=SUM(C{row_idx}:E{row_idx})'
        row_idx += 1

    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20

    path = os.path.join(OUTPUT_DIR, 'customer_inventory_upload.xlsx')
    wb.save(path)
    return path, len(inventory)


def get_sale_row_count():
    """Look up the sale sheet's actual row count so we never truncate."""
    for s in get_workbook_info():
        if s['sheet_id'] == SALE_SHEET_ID:
            return s['row_count']
    return 500


def process_sales():
    row_count = get_sale_row_count()
    data = run_lark_cli([
        "+csv-get", "--url", FEISHU_URL,
        "--sheet-id", SALE_SHEET_ID,
        "--range", f"A1:O{row_count}"
    ])

    headers, rows = parse_csv_rows(data['data']['annotated_csv'])

    non_amazon = []
    for row in rows:
        customer = row.get('Customer', '').strip()
        if 'Amazon' in customer or 'AMAZON' in customer:
            continue
        qty_str = row.get('QTY', '').strip()
        try:
            qty = float(qty_str) if qty_str else 0
        except ValueError:
            qty = 0
        if qty <= 0:
            continue
        non_amazon.append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = '销售明细'
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in non_amazon:
        ws.append([row.get(h, '') for h in headers])

    for i in range(len(headers)):
        ws.column_dimensions[chr(ord('A') + i)].width = 18

    path = os.path.join(OUTPUT_DIR, 'sales_upload_full.xlsx')
    wb.save(path)

    dates = sorted(set(r.get('Ship Date', '') for r in non_amazon if r.get('Ship Date')))
    return path, len(non_amazon), dates


def filter_sales_by_date(cutoff_date_str):
    """cutoff_date_str: 'YYYY/M/D' format, keep only records AFTER this date"""
    from openpyxl import load_workbook
    from datetime import datetime

    cutoff = datetime.strptime(cutoff_date_str.replace('/', '-'), '%Y-%m-%d')

    src_path = os.path.join(OUTPUT_DIR, 'sales_upload_full.xlsx')
    wb_src = load_workbook(src_path)
    ws_src = wb_src.active

    wb = Workbook()
    ws = wb.active
    ws.title = '销售明细'

    headers = [cell.value for cell in ws_src[1]]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    kept = 0
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        ship_date_str = str(row[0]).strip() if row[0] else ''
        if not ship_date_str:
            continue
        try:
            ship_date = datetime.strptime(ship_date_str.split(' ')[0].replace('/', '-'), '%Y-%m-%d')
        except ValueError:
            continue
        if ship_date > cutoff:
            ws.append(list(row))
            kept += 1

    for i in range(len(headers)):
        ws.column_dimensions[chr(ord('A') + i)].width = 18

    path = os.path.join(OUTPUT_DIR, 'sales_upload_filtered.xlsx')
    wb.save(path)
    return path, kept


if __name__ == '__main__':
    action = sys.argv[1] if len(sys.argv) > 1 else 'check'

    if action == 'check':
        sheets = get_workbook_info()
        stock = find_stock_sheet(sheets)
        if stock:
            print(json.dumps({
                'status': 'updated',
                'stock_sheet': stock['sheet_name'],
                'stock_sheet_id': stock['sheet_id'],
                'sheets': [s['sheet_name'] for s in sheets]
            }))
        else:
            print(json.dumps({
                'status': 'not_updated',
                'sheets': [s['sheet_name'] for s in sheets]
            }))

    elif action == 'process':
        sheets = get_workbook_info()
        stock = find_stock_sheet(sheets)
        if not stock:
            print(json.dumps({'error': 'Stock sheet not found'}))
            sys.exit(1)

        inv_path, inv_count = process_inventory(stock)
        sale_path, sale_count, sale_dates = process_sales()

        print(json.dumps({
            'status': 'success',
            'stock_sheet': stock['sheet_name'],
            'inventory': {'path': inv_path, 'sku_count': inv_count},
            'sales': {'path': sale_path, 'record_count': sale_count, 'date_range': sale_dates}
        }))

    elif action == 'filter_sales':
        if len(sys.argv) < 3:
            print("Usage: filter_sales <cutoff_date YYYY/M/D>", file=sys.stderr)
            sys.exit(1)
        path, count = filter_sales_by_date(sys.argv[2])
        print(json.dumps({'path': path, 'kept_records': count}))
